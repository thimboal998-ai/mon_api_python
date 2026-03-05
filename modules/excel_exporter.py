# ============================================
# EXCEL_EXPORTER.PY
# Export Excel multi-feuilles avancé
# ============================================
"""
Exporteur Excel professionnel avec :
- Feuille 1: Données nettoyées
- Feuille 2: Statistiques descriptives
- Feuille 3: Rapport de qualité
- Feuille 4: Graphiques (si disponibles)

Utilise openpyxl pour un contrôle total du formatage.
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage


class ExcelExporter:
    """
    Exporteur Excel avancé avec mise en forme professionnelle
    """
    
    def __init__(self):
        """Initialisation de l'exporteur Excel"""
        self.workbook = None
        self.styles = self._init_styles()
        
        print("📊 ExcelExporter initialisé")
    
    
    def _init_styles(self):
        """
        Initialiser les styles réutilisables
        
        Returns:
            dict: Dictionnaire de styles
        """
        return {
            'header': {
                'font': Font(name='Arial', size=12, bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'title': {
                'font': Font(name='Arial', size=16, bold=True, color='667EEA'),
                'alignment': Alignment(horizontal='left', vertical='center')
            },
            'subtitle': {
                'font': Font(name='Arial', size=12, bold=True, color='764BA2'),
                'alignment': Alignment(horizontal='left', vertical='center')
            },
            'cell': {
                'font': Font(name='Arial', size=10),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                )
            },
            'number': {
                'font': Font(name='Arial', size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'border': Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                ),
                'number_format': '#,##0.00'
            }
        }
    
    
    def export(self, data, stats=None, charts=None, output_path='export.xlsx', metadata=None):
        """
        Exporter les données en Excel multi-feuilles
        
        Args:
            data (DataFrame): Données nettoyées
            stats (dict): Statistiques de nettoyage
            charts (dict): Chemins des graphiques
            output_path (str): Chemin du fichier de sortie
            metadata (dict): Métadonnées supplémentaires
        
        Returns:
            str: Chemin du fichier généré
        """
        print(f"\n📊 Export Excel vers : {output_path}")
        
        # Créer le workbook
        self.workbook = Workbook()
        
        # Supprimer la feuille par défaut
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # 1. Feuille de données nettoyées
        self._create_data_sheet(data)
        
        # 2. Feuille de statistiques descriptives
        if stats:
            self._create_stats_sheet(data, stats)
        
        # 3. Feuille de rapport de qualité (DÉSACTIVÉ DEMANDE UTILISATEUR)
        # if stats:
        #     self._create_quality_report_sheet(stats, metadata)
        
        # 4. Feuille de graphiques (si disponibles)
        if charts:
            self._create_charts_sheet(charts)
        
        # Sauvegarder
        self.workbook.save(output_path)
        
        file_size = os.path.getsize(output_path)
        print(f"   ✅ Excel exporté : {file_size / 1024:.2f} KB")
        print(f"   📑 Feuilles créées : {len(self.workbook.sheetnames)}")
        
        return str(output_path)
    
    
    # ========================================
    # FEUILLE 1: DONNÉES NETTOYÉES
    # ========================================
    
    def _create_data_sheet(self, df):
        """
        Créer la feuille de données nettoyées
        
        Args:
            df (DataFrame): Données à exporter
        """
        print("   📋 Création feuille 'Données Nettoyées'...")
        
        ws = self.workbook.create_sheet('Données Nettoyées', 0)
        
        # En-tête
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['header']['border']
        
        # Données
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                
                # Style selon le type
                if isinstance(value, (int, float)):
                    cell.font = self.styles['number']['font']
                    cell.alignment = self.styles['number']['alignment']
                    cell.number_format = self.styles['number']['number_format']
                else:
                    # Gérer les types complexes (listes, dicts) qui font planter openpyxl
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    
                    cell.value = value
                    cell.font = self.styles['cell']['font']
                    cell.alignment = self.styles['cell']['alignment']
                
                cell.border = self.styles['cell']['border']
        
        # Auto-ajuster les largeurs de colonnes
        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            max_length = len(str(df.columns[col_num - 1]))
            
            # Vérifier la longueur des valeurs
            for row in ws.iter_rows(min_row=2, max_row=min(100, len(df) + 1), min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Limiter la largeur maximale
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Figer la première ligne
        ws.freeze_panes = 'A2'
        
        print(f"      ✅ {len(df)} lignes × {len(df.columns)} colonnes")
    
    
    # ========================================
    # FEUILLE 2: STATISTIQUES DESCRIPTIVES
    # ========================================
    
    def _create_stats_sheet(self, df, stats):
        """
        Créer la feuille de statistiques descriptives
        
        Args:
            df (DataFrame): Données
            stats (dict): Statistiques de nettoyage
        """
        print("   📊 Création feuille 'Statistiques'...")
        
        ws = self.workbook.create_sheet('Statistiques')
        
        # Titre
        ws['A1'] = 'Statistiques Descriptives'
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        ws.merge_cells('A1:F1')
        
        # Statistiques pour colonnes numériques
        numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns
        
        if len(numeric_cols) > 0:
            desc = df[numeric_cols].describe()
            
            # En-têtes
            row = 3
            ws.cell(row=row, column=1, value='Métrique')
            for col_num, col_name in enumerate(numeric_cols, 2):
                cell = ws.cell(row=row, column=col_num, value=col_name)
                cell.font = self.styles['header']['font']
                cell.fill = self.styles['header']['fill']
                cell.alignment = self.styles['header']['alignment']
            
            # Données
            metrics = ['count', 'mean', 'std', 'min', '25%', '50%', '75%', 'max']
            metric_labels = ['Nombre', 'Moyenne', 'Écart-type', 'Minimum', 'Q1', 'Médiane', 'Q3', 'Maximum']
            
            for metric, label in zip(metrics, metric_labels):
                row += 1
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                
                for col_num, col_name in enumerate(numeric_cols, 2):
                    value = desc.loc[metric, col_name]
                    cell = ws.cell(row=row, column=col_num, value=value)
                    cell.number_format = '#,##0.00'
        
        # Ajuster les largeurs
        for col in range(1, len(numeric_cols) + 2):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
        
        print(f"      ✅ Stats pour {len(numeric_cols)} colonnes numériques")
    
    
    # ========================================
    # FEUILLE 3: RAPPORT DE QUALITÉ
    # ========================================
    
    def _create_quality_report_sheet(self, stats, metadata):
        """
        Créer la feuille de rapport de qualité
        
        Args:
            stats (dict): Statistiques de nettoyage
            metadata (dict): Métadonnées
        """
        print("   📄 Création feuille 'Rapport de Qualité'...")
        
        ws = self.workbook.create_sheet('Rapport de Qualité')
        
        s = stats.get('stats', stats)
        quality = stats.get('quality', {})
        
        # Titre
        ws['A1'] = 'Data Quality Report'
        ws['A1'].font = self.styles['title']['font']
        ws.merge_cells('A1:D1')
        
        # Informations générales
        row = 3
        ws.cell(row=row, column=1, value='Date du rapport:')
        ws.cell(row=row, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        
        if metadata:
            row += 1
            ws.cell(row=row, column=1, value='Fichier sur le serveur:')
            ws.cell(row=row, column=2, value=metadata.get('filename', 'N/A'))
        
        # Ligne vide
        row += 2
        
        # ========================================
        # RÉSUMÉ EXÉCUTIF
        # ========================================
        ws.cell(row=row, column=1, value='📊 Résumé Exécutif').font = self.styles['subtitle']['font']
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        lignes_init = s.get('lignes_initiales', 0)
        lignes_final = s.get('lignes_finales', 0)
        
        summary_text = f"L'analyse du fichier a traité {lignes_init} lignes initiales. " \
                       f"Après nettoyage, {lignes_final} lignes ont été conservées pour l'analyse."
        
        ws.cell(row=row, column=1, value=summary_text)
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 40
        row += 2

        # ========================================
        # INDICATEURS CLÉS
        # ========================================
        ws.cell(row=row, column=1, value='🔑 Indicateurs Clés').font = self.styles['subtitle']['font']
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        kpis = [
            ('Lignes initiales', s.get('lignes_initiales', 'N/A')),
            ('Lignes finales', s.get('lignes_finales', 'N/A')),
            ('Colonnes', s.get('colonnes_finales', 'N/A')),
            ('Valeurs manquantes traitées', s.get('valeurs_manquantes_traitees', 0)),
            ('Outliers traités', s.get('lignes_outliers_traitees', 0)),
            ('Doublons supprimés', s.get('doublons_supprimes', 0)),
            ('Score de qualité', f"{quality.get('score', 0)}%")
        ]
        
        for label, value in kpis:
            row += 1
            cell_label = ws.cell(row=row, column=1, value=label)
            cell_label.font = Font(bold=True)
            
            cell_value = ws.cell(row=row, column=2, value=value)
            
            # Couleur selon la métrique
            if 'Score' in label:
                score = quality.get('score', 0)
                if score >= 90:
                    cell_value.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif score >= 70:
                    cell_value.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell_value.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Ligne vide
        row += 2
        
        # Transformations
        if stats.get('transformation_history'):
            ws.cell(row=row, column=1, value='🔄 Historique des Transformations').font = self.styles['subtitle']['font']
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            for i, transform in enumerate(stats['transformation_history'], 1):
                row += 1
                ws.cell(row=row, column=1, value=f"{i}.")
                ws.cell(row=row, column=2, value=transform.get('step', 'N/A'))
                ws.cell(row=row, column=3, value=f"{transform.get('count', 0)} éléments")
        
        row += 2
        
        # ========================================
        # RECOMMANDATIONS
        # ========================================
        ws.cell(row=row, column=1, value='💡 Recommandations').font = self.styles['subtitle']['font']
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        recs = []
        if s.get('valeurs_manquantes_traitees', 0) > 0:
            recs.append("✅ Les valeurs manquantes ont été traitées. Vérifiez si la méthode (moyenne/médiane) convient à votre cas d'usage.")
        if s.get('lignes_outliers_traitees', 0) > 0:
            recs.append("✅ Des valeurs aberrantes (outliers) ont été détectées et traitées.")
        if s.get('doublons_supprimes', 0) > 0:
            recs.append("✅ Les doublons exacts ont été supprimés pour éviter les biais.")
            
        recs.append("📊 N'oubliez pas de consulter l'onglet 'Graphiques' pour visualiser la distribution des données.")
        
        for rec in recs:
            row += 1
            ws.cell(row=row, column=1, value=f"• {rec}")
            ws.merge_cells(f'A{row}:F{row}')
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        
        print(f"      ✅ Rapport créé avec {len(kpis)} KPIs et {len(recs)} recommandations")
    
    
    # ========================================
    # FEUILLE 4: GRAPHIQUES
    # ========================================
    
    def _create_charts_sheet(self, charts):
        """
        Créer la feuille de graphiques
        
        Args:
            charts (dict): Chemins des graphiques
        """
        print("   📈 Création feuille 'Graphiques'...")
        
        ws = self.workbook.create_sheet('Graphiques')
        
        # Titre
        ws['A1'] = 'Visualisations'
        ws['A1'].font = self.styles['title']['font']
        ws.merge_cells('A1:H1')
        
        row = 3
        
        # Convertir charts en liste
        if isinstance(charts, dict):
            chart_paths = list(charts.values())
        else:
            chart_paths = charts
        
        # Insérer les images
        for i, chart_path in enumerate(chart_paths):
            # Convertir chemin web en chemin système
            if isinstance(chart_path, str):
                chart_path = chart_path.replace('/static/', 'static/')
            
            if os.path.exists(chart_path):
                try:
                    # Créer l'objet image
                    img = XLImage(chart_path)
                    
                    # Redimensionner pour tenir dans la feuille
                    # Largeur max: 8 colonnes (~640 pixels)
                    max_width = 640
                    max_height = 480
                    
                    # Calculer le ratio
                    if img.width > max_width or img.height > max_height:
                        ratio = min(max_width / img.width, max_height / img.height)
                        img.width = int(img.width * ratio)
                        img.height = int(img.height * ratio)
                    
                    # Positionner l'image
                    ws.add_image(img, f'A{row}')
                    
                    # Passer à la ligne suivante (estimation: 1 ligne = 20 pixels)
                    row += int(img.height / 20) + 2
                    
                    print(f"      ✅ Graphique {i+1} inséré")
                
                except Exception as e:
                    print(f"      ⚠️  Erreur insertion image {i+1}: {str(e)}")
        
        print(f"      ✅ {len(chart_paths)} graphique(s) traité(s)")


print("✅ ExcelExporter chargé avec succès")